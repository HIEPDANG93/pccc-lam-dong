#!/usr/bin/env python3
"""Convert the customer's split Excel workbook into Google Sheet import CSVs."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


BUSINESS_HEADERS = [
    "id",
    "tenCoSo",
    "bangHieu",
    "nganhNghe",
    "tinhCu",
    "xaPhuong",
    "diaChi",
    "diaDiemKinhDoanh",
    "nguoiDungDau",
    "soDienThoai",
    "soGpkd",
    "ngayCapPccc",
    "soGiayAntt",
    "soPhong",
    "soGiuong",
    "soLaoDong",
    "latitude",
    "longitude",
    "googleMapsUrl",
    "trangThai",
    "updatedAt",
]

SHEET_INDUSTRY = {
    "LuuTru_C06": "Lưu trú",
    "LuuTru_XoaBop_C06": "Lưu trú; Xoa bóp",
    "Luutru_Karaoke_C06": "Lưu trú; Karaoke",
    "Luutru_Karaoke_Xoabop_C06": "Lưu trú; Karaoke; Xoa bóp",
    "QuanTrang_C06": "Quân trang",
}

FIELD_MAP = {
    "TÊN CƠ SỞ": "tenCoSo",
    "BẢNG HIỆU": "bangHieu",
    "LOẠI HÌNH KINH DOANH": "nganhNghe",
    "NGÀNH NGHỀ": "nganhNghe",
    "TỈNH (CŨ)": "tinhCu",
    "XÃ, PHƯỜNG": "xaPhuong",
    "ĐỊA CHỈ": "diaChi",
    "ĐỊA ĐIỂM KINH DOANH": "diaDiemKinhDoanh",
    "NGƯỜI ĐỨNG ĐẦU": "nguoiDungDau",
    "SỐ ĐIỆN THOẠI": "soDienThoai",
    "SỐ GPKD": "soGpkd",
    "NGÀY CẤP PCCC": "ngayCapPccc",
    "SỐ GIẤY CHỨNG NHẬN ANTT": "soGiayAntt",
    "SỐ PHÒNG": "soPhong",
    "SỐ GIƯỜNG": "soGiuong",
    "TỔNG SỐ NGƯỜI LAO ĐỘNG": "soLaoDong",
    "ĐỊNH VỊ": "googleMapsUrl",
}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "0":
        return ""
    return re.sub(r"\s+", " ", text)


def parse_coordinate(value: str) -> tuple[str, str, str]:
    text = clean(value)
    numbers = re.findall(r"-?\d+(?:\.\d+)?", text)
    if len(numbers) >= 2:
        return numbers[0], numbers[1], text
    return "", "", text


def merge_industries(*values: str) -> str:
    seen: list[str] = []
    for value in values:
        for part in re.split(r"[,;]+", clean(value)):
            item = part.strip()
            if item and item.lower() not in [entry.lower() for entry in seen]:
                seen.append(item)
    return "; ".join(seen)


def convert(source: Path, output_dir: Path) -> Path:
    wb = load_workbook(source, data_only=True, read_only=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, str]] = []
    seen_keys: set[str] = set()

    for sheet_name, default_industry in SHEET_INDUSTRY.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        for raw in ws.iter_rows(min_row=2, values_only=True):
            source_row = {headers[index]: clean(value) for index, value in enumerate(raw[: len(headers)])}
            if not source_row.get("TÊN CƠ SỞ"):
                continue
            key = clean(source_row.get("TÊN CƠ SỞ", "") + "|" + source_row.get("ĐỊA ĐIỂM KINH DOANH", "")).lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)

            item = {header: "" for header in BUSINESS_HEADERS}
            for excel_name, field_name in FIELD_MAP.items():
                if excel_name in source_row:
                    item[field_name] = source_row[excel_name]

            lat, lng, maps_url = parse_coordinate(item["googleMapsUrl"])
            item["latitude"] = lat
            item["longitude"] = lng
            item["googleMapsUrl"] = maps_url
            item["nganhNghe"] = merge_industries(default_industry, item["nganhNghe"])
            item["trangThai"] = "Đang hoạt động"
            rows.append(item)

    for index, item in enumerate(rows, start=1):
        item["id"] = f"CS-{index:04d}"

    output_file = output_dir / "DoanhNghiep.csv"
    with output_file.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=BUSINESS_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    return output_file


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: import_excel_to_csv.py <source.xlsx> [output_dir]", file=sys.stderr)
        return 2
    source = Path(sys.argv[1]).expanduser().resolve()
    output_dir = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else Path("data/import").resolve()
    output = convert(source, output_dir)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
